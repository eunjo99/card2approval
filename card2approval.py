# -*- coding: utf-8 -*-
"""
카드내역 원본 엑셀 -> 회사 포맷(멀티 시트) + 보고용 테이블 생성
- CLI와 Streamlit 모두 호환
- 필요 패키지: pandas, openpyxl (xls 지원 시 xlrd==2.0.1)
"""

import io
import os
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# -------------------------
# 설정 상수
# -------------------------
SOURCE_COLS = [
    "카드번호","승인일자","승인시간","승인금액(원화)","승인금액(외화)",
    "공급가액(원화)","부가세","외화거래일환율","외화거래국가코드",
    "가맹점사업자번호","가맹점명","가맹점업종명"
]

OUTPUT_ORDER = [
    "카드번호","승인일자","승인시간","승인금액(원화)","승인금액(외화)",
    "공급가액(원화)","부가세","외화거래일환율","외화거래국가코드",
    "가맹점사업자번호","가맹점명","가맹점업종명","적요","비용 구분"
]

CURRENCY_COLS = ["승인금액(원화)","승인금액(외화)","공급가액(원화)","부가세"]


# -------------------------
# 유틸 함수
# -------------------------
def _is_file_like(x) -> bool:
    return hasattr(x, "read")


def read_raw(source, sheet=0):
    """
    원본 엑셀 읽기 (경로 또는 파일객체 둘 다 지원)
    """
    df = pd.read_excel(source, sheet_name=sheet, dtype=str)
    cols_present = [c for c in SOURCE_COLS if c in df.columns]
    df = df[cols_present].copy()

    # 날짜 포맷
    if "승인일자" in df.columns:
        try:
            df["승인일자"] = pd.to_datetime(df["승인일자"], errors="coerce").dt.strftime("%Y.%m.%d")
        except Exception:
            pass

    # 금액류 숫자화(콤마 제거) - 원화 합계용
    for col in ["승인금액(원화)","공급가액(원화)","부가세"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(",",""), errors="coerce").fillna(0).astype(int)

    return df


def load_mapping(path: str):
    """
    로컬/경로 CSV에서 매핑 읽기 (title 없이도 OK)
    기대 컬럼: card_number_masked, employee_name, site
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"매핑 CSV 경로를 찾을 수 없습니다: {path}")
    df = pd.read_csv(path, dtype=str).fillna("")
    expected_min = {"card_number_masked","employee_name","site"}
    missing = expected_min - set(df.columns)
    if missing:
        raise ValueError(f"매핑 파일에 누락된 컬럼: {missing}")
    # title이 있어도 무시, 없어도 문제 없음
    return df


def load_mapping_from_upload(file_like):
    """
    업로드된 매핑 파일(CSV 또는 XLSX) 읽기 (title 없이도 OK)
    기대 컬럼: card_number_masked, employee_name, site
    """
    name = getattr(file_like, "name", "").lower()
    if name.endswith(".csv"):
        df = pd.read_csv(file_like, dtype=str).fillna("")
    elif name.endswith(".xlsx"):
        df = pd.read_excel(file_like, dtype=str, engine="openpyxl").fillna("")
    else:
        raise ValueError("매핑 파일은 CSV 또는 XLSX만 지원합니다.")

    expected_min = {"card_number_masked","employee_name","site"}
    missing = expected_min - set(df.columns)
    if missing:
        raise ValueError(f"매핑 파일에 누락된 컬럼: {missing}")
    # title 있으면 무시
    return df

def _apply_currency_format(ws):
    """
    표 형태 시트(전체/판교/대전)에 금액 콤마 서식 적용
    """
    if ws.max_row < 2:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def _build_workbook(df_raw, mapping, month_label: str) -> Workbook:
    """
    공통 워크북 생성 로직 (전체/판교/대전/보고용)
    """
    # 매핑 병합
    df = df_raw.merge(mapping, left_on="카드번호", right_on="card_number_masked", how="left")

    # ✅ 한글 컬럼 생성 (시트1/2/3에서 사용)
    df["직원명"] = df.get("employee_name", "")
    df["사업장"] = df.get("site", "")
    # '전체' 시트용 DF
    whole_cols = SOURCE_COLS + ["직원명","사업장"]
    whole_cols = [c for c in whole_cols if c in df.columns]
    df_whole = df[whole_cols].copy()

    wb = Workbook()

    # 시트1: 전체
    ws_all = wb.active
    ws_all.title = "전체"
    for r in dataframe_to_rows(df_whole, index=False, header=True):
        ws_all.append(r)
    _apply_currency_format(ws_all)

    # 시트2/3: 판교/대전
    def add_filtered(name):
        ws = wb.create_sheet(title=name)
        dfx = df_whole[df_whole["사업장"] == name].copy() if "사업장" in df_whole.columns else df_whole.iloc[0:0].copy()
        for r in dataframe_to_rows(dfx, index=False, header=True):
            ws.append(r)
        _apply_currency_format(ws)
        return ws

    add_filtered("판교")
    add_filtered("대전")

    # 시트4: 보고용
    ws = wb.create_sheet(title="보고용")
    title = f"{month_label} 법인카드 이용내역" if month_label else "법인카드 이용내역"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    row_cursor = 3
    grand_total = 0

    # 정렬
    sort_cols = [c for c in ["site","employee_name","카드번호","승인일자","승인시간"] if c in df.columns]
    df_sorted = df.sort_values(sort_cols).reset_index(drop=True)

    for site in [s for s in ["판교","대전"] if ("site" in df_sorted.columns and s in df_sorted["site"].unique().tolist())]:
        site_df = df_sorted[df_sorted["site"] == site].copy()
        site_total = 0

        # 사이트 헤더
        ws.cell(row=row_cursor, column=1, value=f"{site} 내역")
        ws.cell(row=row_cursor, column=1).font = Font(bold=True)
        row_cursor += 1

        for (emp, _), g in site_df.groupby(["employee_name","카드번호"], dropna=False):
            # 헤더 행
            for ci, val in enumerate(OUTPUT_ORDER, start=1):
                c = ws.cell(row=row_cursor, column=ci, value=val)
                c.font = Font(bold=True)
                c.fill = header_fill
                c.border = border_all
                c.alignment = Alignment(horizontal="center")
            header_row = row_cursor
            row_cursor += 1

            # 데이터 행(적요/비용 구분 비움)
            g = g.copy()
            g["적요"] = ""
            g["비용 구분"] = ""
            g = g[[c for c in OUTPUT_ORDER if c in g.columns]]

            start_data_row = row_cursor
            for _, r in g.iterrows():
                for ci, col in enumerate(OUTPUT_ORDER, start=1):
                    val = r.get(col, "")
                    cell = ws.cell(row=row_cursor, column=ci, value=val)
                    cell.border = border_all
                row_cursor += 1
            end_data_row = row_cursor - 1

            # 금액 서식(내역)
            for col_letter, col_name in zip(list("ABCDEFGHIJKLMN"), OUTPUT_ORDER):
                if col_name in CURRENCY_COLS:
                    for rr in range(start_data_row, end_data_row+1):
                        ws[f"{col_letter}{rr}"].number_format = "#,##0"

            # 직원 합계
            emp_sum = int(g["승인금액(원화)"].sum()) if "승인금액(원화)" in g.columns else 0
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
            total_label = ws.cell(row=row_cursor, column=1, value=f"{emp} 합계")
            total_label.font = Font(bold=True)
            total_label.alignment = Alignment(horizontal="center")
            total_amt = ws.cell(row=row_cursor, column=4, value=emp_sum)
            total_amt.font = Font(bold=True)
            total_amt.number_format = "#,##0"

            # 테두리(헤더~합계까지)
            for cc in range(1, 15):
                ws.cell(row=header_row, column=cc).border = border_all
            for rr in range(start_data_row, row_cursor+1):
                for cc in range(1, 15):
                    ws.cell(row=rr, column=cc).border = border_all

            row_cursor += 1          # 합계 줄
            row_cursor += 1          # 공백 1행
            site_total += emp_sum

        # 사이트 합계
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
        lab = ws.cell(row=row_cursor, column=1, value=f"{site} 합계")
        lab.font = Font(bold=True)
        lab.alignment = Alignment(horizontal="center")
        amt = ws.cell(row=row_cursor, column=4, value=site_total)
        amt.font = Font(bold=True)
        amt.number_format = "#,##0"
        for cc in range(1, 5):
            ws.cell(row=row_cursor, column=cc).border = border_all

        row_cursor += 2
        grand_total += site_total

    # 전체 합계
    ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
    lab = ws.cell(row=row_cursor, column=1, value="전체 합계")
    lab.font = Font(bold=True)
    lab.alignment = Alignment(horizontal="center")
    amt = ws.cell(row=row_cursor, column=4, value=grand_total)
    amt.font = Font(bold=True)
    amt.number_format = "#,##0"
    for cc in range(1, 5):
        ws.cell(row=row_cursor, column=cc).border = border_all

    # 열 너비
    widths = [18,12,10,14,14,14,10,16,16,18,24,18,18,12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    return wb


# -------------------------
# 공개 API (Streamlit/CLI 공용)
# -------------------------
def build_excel_bytes(df_raw, mapping, month_label: str) -> io.BytesIO:
    """
    Streamlit에서 다운로드용으로 사용하는 in-memory 엑셀 생성
    """
    wb = _build_workbook(df_raw, mapping, month_label)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_multi_sheet(df_raw, mapping, month_label: str, out_path: str):
    """
    CLI에서 파일을 직접 저장하는 버전
    """
    wb = _build_workbook(df_raw, mapping, month_label)
    wb.save(out_path)


# -------------------------
# CLI 진입점 (선택 사용)
# -------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", required=True, help="원본 엑셀 경로(.xlsx 권장)")
    ap.add_argument("--sheet", default=0, help="시트명 또는 인덱스")
    ap.add_argument("--mapping", required=True, help="card_employee_mapping.csv 경로")
    ap.add_argument("--out", required=True, help="출력 엑셀 경로(.xlsx)")
    ap.add_argument("--month", default="", help='예: "7월" (보고용 시트 제목에 사용)')
    args = ap.parse_args()

    df_raw = read_raw(args.raw, args.sheet)
    mapping = load_mapping(args.mapping)
    build_multi_sheet(df_raw, mapping, args.month, args.out)
    print(f"[OK] 생성: {args.out}")


if __name__ == "__main__":
    main()
